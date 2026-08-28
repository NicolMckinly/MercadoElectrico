"""
procesar_limitacion.py

Lee los dos archivos Excel de limitación de suministro descargados de XM
(cada uno con hojas Inf_Detallado, Ultimos_Iniciados, Ultimos_Cancelados),
filtra la información de los últimos 8 días y elimina duplicados.

No descarga nada ni conoce a Teams: solo transforma los .xlsx en una
estructura de datos lista para armar el mensaje.
"""

import openpyxl
from datetime import datetime, timedelta


def _normalizar_encabezado(valor):
    """Convierte un encabezado de columna a texto limpio y comparable."""
    if valor is None:
        return ""
    return str(valor).strip().lower()


def _encontrar_fila_encabezado(ws, columna_clave="sigla empresa"):
    """
    Busca la fila donde está el encabezado real de la tabla (la que
    contiene 'Sigla empresa'), porque las primeras filas de estas hojas
    son títulos y notas que varían de tamaño entre archivos.
    Devuelve el número de fila (1-indexado) o None si no la encuentra.
    """
    for fila in ws.iter_rows(min_row=1, max_row=20):
        for celda in fila:
            if _normalizar_encabezado(celda.value) == columna_clave:
                return celda.row
    return None


def _mapear_columnas(ws, fila_encabezado):
    """Devuelve un diccionario {nombre_columna_normalizado: índice_columna}."""
    mapa = {}
    for celda in ws[fila_encabezado]:
        nombre = _normalizar_encabezado(celda.value)
        if nombre:
            mapa[nombre] = celda.column
    return mapa


def _leer_filas_datos(ws, fila_encabezado):
    """
    Devuelve la lista de filas de datos (como tuplas de valores por fila),
    empezando justo debajo del encabezado y hasta que aparezcan varias
    filas vacías seguidas (fin de la tabla).
    """
    filas = []
    vacias_seguidas = 0
    for fila in ws.iter_rows(min_row=fila_encabezado + 1):
        valores = [c.value for c in fila]
        if all(v is None or str(v).strip() == "" for v in valores):
            vacias_seguidas += 1
            if vacias_seguidas >= 3:
                break
            continue
        vacias_seguidas = 0
        filas.append(fila)
    return filas


def _valor_columna(fila, mapa_columnas, nombre_columna, fila_encabezado_num):
    """Obtiene el valor de una columna por nombre normalizado, o None si no existe."""
    idx = mapa_columnas.get(nombre_columna)
    if idx is None:
        return None
    # fila es una tupla de celdas empezando en la primera columna de la hoja
    for celda in fila:
        if celda.column == idx:
            return celda.value
    return None


def _a_fecha(valor):
    """Convierte el valor de una celda de fecha (datetime o texto) a date."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if hasattr(valor, "year") and hasattr(valor, "month"):  # objeto date
        return valor
    try:
        return datetime.strptime(str(valor).strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def _extraer_seccion(ws, columna_fecha_filtro, fecha_corte, campo_nombre="empresa"):
    """
    Extrae de una hoja (Ultimos_Iniciados o Ultimos_Cancelados) las filas
    dentro de los últimos 8 días, deduplicadas por (Empresa, Actividad).

    columna_fecha_filtro: nombre normalizado de la columna de fecha a usar
                           ('fecha vencimiento' o 'fecha cancelación')
    fecha_corte: fecha (date) más antigua permitida (inclusive)

    Devuelve una lista de tuplas (actividad, nombre_empresa) sin duplicados,
    en el orden en que aparecen.
    """
    fila_encabezado = _encontrar_fila_encabezado(ws)
    if fila_encabezado is None:
        return []

    mapa = _mapear_columnas(ws, fila_encabezado)
    filas_datos = _leer_filas_datos(ws, fila_encabezado)

    vistos = set()
    resultado = []

    for fila in filas_datos:
        fecha_valor = _valor_columna(fila, mapa, columna_fecha_filtro, fila_encabezado)
        fecha = _a_fecha(fecha_valor)
        if fecha is None or fecha < fecha_corte:
            continue

        empresa = _valor_columna(fila, mapa, "empresa", fila_encabezado)
        actividad = _valor_columna(fila, mapa, "actividad", fila_encabezado)

        if empresa is None:
            continue

        empresa = str(empresa).strip()
        actividad = str(actividad).strip().title() if actividad else ""

        clave = (empresa.upper(), actividad.upper())
        if clave in vistos:
            continue
        vistos.add(clave)
        resultado.append((actividad, empresa))

    return resultado


def procesar_archivo(ruta_excel, fecha_ejecucion=None, dias_atras=8):
    """
    Procesa un archivo de limitación de suministro (cualquiera de los dos
    tipos: 'corte a usuarios' o 'en bolsa') y devuelve:

        {
            "iniciados": [(actividad, empresa), ...],
            "cancelados": [(actividad, empresa), ...],
        }

    fecha_ejecucion: fecha de referencia (por defecto, hoy). Se filtra
                      todo lo que esté entre (fecha_ejecucion - dias_atras)
                      y fecha_ejecucion, inclusive.
    """
    if fecha_ejecucion is None:
        fecha_ejecucion = datetime.now().date()
    fecha_corte = fecha_ejecucion - timedelta(days=dias_atras)

    wb = openpyxl.load_workbook(ruta_excel, data_only=True)

    ws_iniciados = wb["Ultimos_Iniciados"]
    ws_cancelados = wb["Ultimos_Cancelados"]

    iniciados = _extraer_seccion(
        ws_iniciados, columna_fecha_filtro="fecha vencimiento", fecha_corte=fecha_corte
    )
    cancelados = _extraer_seccion(
        ws_cancelados, columna_fecha_filtro="fecha cancelación", fecha_corte=fecha_corte
    )

    return {"iniciados": iniciados, "cancelados": cancelados}


if __name__ == "__main__":
    # Prueba rápida manual: ajusta las rutas y corre "python procesar_limitacion.py"
    import sys

    ruta = sys.argv[1] if len(sys.argv) > 1 else "archivo.xlsx"
    resultado = procesar_archivo(ruta)
    print("ÚLTIMOS INICIADOS:")
    for actividad, empresa in resultado["iniciados"]:
        print(f"  {actividad} | {empresa}")
    print("\nÚLTIMOS CANCELADOS:")
    for actividad, empresa in resultado["cancelados"]:
        print(f"  {actividad} | {empresa}")
